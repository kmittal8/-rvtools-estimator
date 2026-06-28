import streamlit as st
import pandas as pd
import requests
from io import BytesIO

st.set_page_config(page_title="OCI Migration Estimator", page_icon="☁️", layout="wide")

st.markdown("""
<style>
/* ── Base ── */
html, body { background: #eef0f4 !important; }
[data-testid="stAppViewContainer"] { background: #eef0f4 !important; }
[data-testid="stHeader"],[data-testid="stToolbar"],[data-testid="stDecoration"] { display:none !important; }
.main .block-container { padding: 0 0 40px 0 !important; max-width:100% !important; }

/* ── Metric cards ── */
[data-testid="metric-container"] {
    background: #ffffff !important;
    border: 1px solid #dde2e8 !important;
    border-radius: 6px !important;
    padding: 18px 20px 14px !important;
    box-shadow: 0 1px 3px rgba(0,0,0,0.08) !important;
}
[data-testid="stMetricValue"] {
    font-size: 1.65rem !important; font-weight: 700 !important; color: #1c1c1e !important;
}
[data-testid="stMetricLabel"] {
    font-size: 0.72rem !important; font-weight: 600 !important;
    color: #6e7681 !important; text-transform: uppercase !important; letter-spacing: 0.05em !important;
}

/* ── Tabs ── */
[data-testid="stTabs"] button {
    font-size: 0.87rem !important; font-weight: 500 !important;
    color: #555 !important; padding: 12px 22px !important; border-radius: 0 !important;
}
[data-testid="stTabs"] button[aria-selected="true"] {
    color: #1a5c31 !important; font-weight: 700 !important;
    border-bottom: 3px solid #1a5c31 !important;
}

/* ── Section headings ── */
h3 {
    font-size: 0.72rem !important; font-weight: 700 !important;
    color: #6e7681 !important; text-transform: uppercase !important;
    letter-spacing: 0.09em !important; margin: 28px 0 12px !important;
}

/* ── Dividers ── */
hr { border-color: #dde2e8 !important; }

/* ── DataFrames ── */
[data-testid="stDataFrame"] {
    border: 1px solid #dde2e8 !important; border-radius: 6px !important;
    background: #fff !important; box-shadow: 0 1px 3px rgba(0,0,0,0.05) !important;
}

/* ── Inputs ── */
[data-testid="stSelectbox"] > div > div,
[data-testid="stNumberInput"] input {
    background: #fff !important; border: 1px solid #c8cdd5 !important;
    border-radius: 4px !important; color: #1c1c1e !important;
}

/* ── File uploader ── */
[data-testid="stFileUploader"] {
    background: #fff !important; border: 2px dashed #c0c8d2 !important;
    border-radius: 6px !important;
}

/* ── Alerts ── */
[data-testid="stAlert"] {
    border-radius: 4px !important;
    border-left: 4px solid #e0a800 !important;
    background: #fffcf0 !important;
}

/* ── Expanders ── */
[data-testid="stExpander"] {
    background: #fff !important; border: 1px solid #dde2e8 !important;
    border-radius: 6px !important; box-shadow: 0 1px 2px rgba(0,0,0,0.04) !important;
}

/* ── Download button ── */
[data-testid="stDownloadButton"] button {
    background: #1a5c31 !important; color: #fff !important;
    border: none !important; border-radius: 4px !important;
    font-weight: 600 !important; padding: 10px 22px !important;
    box-shadow: 0 1px 4px rgba(0,0,0,0.15) !important;
}
[data-testid="stDownloadButton"] button:hover { background: #144d28 !important; }

/* ── Captions ── */
[data-testid="stCaptionContainer"] p { color: #6e7681 !important; font-size: 0.78rem !important; }

/* ── Content padding inside tabs ── */
[data-testid="stTabsContent"] { padding: 24px 32px !important; }
</style>
""", unsafe_allow_html=True)

# ── Header bar ─────────────────────────────────────────────
st.markdown("""
<div style="background:#1a5c31;padding:14px 32px 13px;
            display:flex;align-items:center;justify-content:space-between;
            margin:-1px -1px 0 -1px;">
  <div style="display:flex;align-items:center;gap:12px;">
    <div style="width:36px;height:36px;border-radius:8px;
                background:rgba(255,255,255,0.15);
                display:flex;align-items:center;justify-content:center;font-size:1.2rem;">☁️</div>
    <div>
      <div style="font-size:1.05rem;font-weight:700;color:#fff;line-height:1.2;
                  font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;">
        OCI Migration Estimator</div>
      <div style="font-size:0.71rem;color:rgba(255,255,255,0.6);margin-top:2px;letter-spacing:0.02em;">
        VMware → OCI &nbsp;·&nbsp; Live pricing API &nbsp;·&nbsp; Upload RVTools export to begin</div>
    </div>
  </div>
  <div style="font-size:0.68rem;color:rgba(255,255,255,0.45);letter-spacing:0.02em;">
    Not an official Oracle product</div>
</div>
""", unsafe_allow_html=True)

st.markdown("""
<div style="background:#fff;border-bottom:1px solid #dde3ea;padding:10px 32px;
            display:flex;align-items:center;gap:10px;">
  <span style="color:#f5a623;font-size:1rem;">⚠️</span>
  <span style="font-size:0.8rem;color:#555;">
    Always validate pricing with the official
    <a href="https://www.oracle.com/anz/cloud/costestimator.html" target="_blank"
       style="color:#2d6a3f;font-weight:600;">OCI Cost Estimator</a>
    before sharing with customers.
  </span>
</div>
<div style="background:#fff;padding:16px 32px;border-bottom:1px solid #dde3ea;">
""", unsafe_allow_html=True)
uploaded_file = st.file_uploader("Upload RVTools Excel export (.xlsx)", type=["xlsx"], label_visibility="collapsed")
st.markdown("</div>", unsafe_allow_html=True)

VMWARE_NAME_PREFIXES = ('vcls-', 'z-vrah-', 'z-vra-')
HOURS_PER_MONTH = 744
BALANCED_VPU = 10  # OCI Balanced performance tier = 10 VPU

# OCI SKUs
SKU_E4_OCPU   = 'B93113'
SKU_E4_MEM    = 'B93114'
SKU_BLOCK_STG = 'B91961'
SKU_BLOCK_VPU = 'B91962'
SKU_OBJ_STG   = 'B96484'

@st.cache_data(ttl=3600)
def fetch_oci_prices(currency):
    url = f"https://apexapps.oracle.com/pls/apex/cetools/api/v1/products/?currencyCode={currency}"
    r = requests.get(url, timeout=10)
    r.raise_for_status()
    items = r.json()['items']
    return {i['partNumber']: i['currencyCodeLocalizations'][0]['prices'][0]['value'] for i in items}

def os_family(name):
    n = str(name).lower()
    if 'windows' in n:
        return 'Windows'
    elif any(x in n for x in ['linux', 'debian', 'ubuntu', 'suse', 'red hat', 'centos', 'photon']):
        return 'Linux'
    else:
        return 'Other'

DISCLAIMER = (
    "Disclaimer:  This sample quote is provided solely for evaluation purposes and is intended to further "
    "discussions between you and Oracle.  This sample quote is not eligible for acceptance by you and is not "
    "a binding contract between you and Oracle for the services specified.  If you would like to purchase the "
    "services specified in this sample quote, please request that Oracle issue you a formal quote (which may "
    "include an OMA or a CSA if you do not already have an appropriate agreement in place with Oracle) for "
    "your acceptance and execution.  Your formal quote will be effective only upon Oracle's acceptance of the "
    "formal quote (and the OMA or CSA, if required)."
)

SKU_DESCRIPTIONS = {
    SKU_E4_OCPU:   ("Compute - Standard - E4 - OCPU (OCPU Per Hour)\nCapacity Type: On-Demand",   "OCPU Per Hour"),
    SKU_E4_MEM:    ("Compute - Standard - E4 - Memory (Gigabyte Per Hour)\nCapacity Type: On-Demand", "Gigabyte Per Hour"),
    SKU_BLOCK_STG: ("Storage - Block Volume - Storage (Gigabyte Storage Capacity Per Month)",      "Gigabyte Storage Capacity Per Month"),
    SKU_BLOCK_VPU: ("Storage - Block Volume - Performance Units (Performance Units Per Gigabyte Per Month)", "Performance Units Per Gigabyte Per Month"),
}

def to_excel(df):
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Migration BOM')
    return buf.getvalue()

def to_oci_excel(bom_df, prices, currency, shape, rw_params=None, obj_params=None, custom_hours=False, free_tier_discount=0.0):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from datetime import date

    wb = Workbook()
    ws = wb.active
    ws.title = "Migration BOM"

    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 60
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 16
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 15

    bold        = Font(bold=True)
    bold_lg     = Font(bold=True, size=12)
    red_bold    = Font(bold=True, color="FF0000")
    group_font  = Font(bold=True)
    light_blue  = PatternFill("solid", fgColor="DCE6F1")
    light_green = PatternFill("solid", fgColor="E2EFDA")
    light_grey  = PatternFill("solid", fgColor="F2F2F2")
    wrap        = Alignment(wrap_text=True, vertical="top")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'),  bottom=Side(style='thin')
    )

    ocpu_rate   = prices[SKU_E4_OCPU]
    mem_rate    = prices[SKU_E4_MEM]
    stg_rate    = prices[SKU_BLOCK_STG]
    vpu_rate    = prices[SKU_BLOCK_VPU]
    win_os_rate = prices.get('B88318', 0)

    # Aggregate by OS family + BYOL
    win_byol  = bom_df[(bom_df['OS Family'] == 'Windows') & (bom_df['BYOL'] == True)]
    win_li    = bom_df[(bom_df['OS Family'] == 'Windows') & (bom_df['BYOL'] == False)]
    linux_df  = bom_df[bom_df['OS Family'] != 'Windows']

    def agg(df):
        return {
            'vms':    len(df),
            'ocpus':  int(df['OCPUs'].sum()),
            'mem':    float(df['Memory GB'].sum()),
            'stg':    float(df['Provisioned GB'].sum()),
            'df':     df,
        }

    def group_cost(a, include_win_os=False):
        if custom_hours and 'Compute/mo' in a['df'].columns:
            compute = float(a['df']['Compute/mo'].sum())
            storage = float(a['df']['Storage/mo'].sum())
            wos     = float(a['df']['WinOS/mo'].sum()) if include_win_os else 0
        else:
            compute = (a['ocpus'] * ocpu_rate + a['mem'] * mem_rate) * HOURS_PER_MONTH
            storage = a['stg'] * stg_rate + (a['stg'] * BALANCED_VPU) * vpu_rate
            wos     = (a['ocpus'] * win_os_rate * HOURS_PER_MONTH) if include_win_os else 0
        return round(compute, 2), round(storage, 2), round(wos, 2), round(compute + storage + wos, 2)

    row = 1
    today = date.today().strftime("%m/%d/%Y")
    ws.cell(row=row, column=1, value=f"Oracle Investment Proposal (as of {today})").font = bold_lg
    row += 1
    ws.cell(row=row, column=1, value="Reference label: Migration Estimate")
    row += 1
    ws.cell(row=row, column=1, value=f"Currency: {currency}")
    row += 1
    ws.cell(row=row, column=1, value=f"Shape: {shape}  |  Capacity Type: On-Demand  |  Storage: Balanced (10 VPU)")
    row += 1
    hrs_note = "  |  Note: Some VMs use custom Hrs/Month — compute cost reflects actual hours, storage is full month." if custom_hours else ""
    ws.cell(row=row, column=1, value=f"Realm: PUBLIC  |  Service Type: IAAS{hrs_note}")
    row += 2

    # Column headers
    headers = ["Part", "Description", "VMs", "Total OCPUs", "Total Mem GB", "Unit Price", "Monthly Cost", "Notes"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = bold
        c.fill = light_blue
    row += 1

    grand_total = 0.0

    def write_group(label, a, include_win_os, fill):
        nonlocal row, grand_total
        compute, storage, wos, total = group_cost(a, include_win_os)
        grand_total += total
        vpu_qty = int(a['stg'] * BALANCED_VPU)

        # Group header
        hc = ws.cell(row=row, column=2, value=label)
        hc.font = group_font
        hc.fill = fill
        ws.cell(row=row, column=3, value=a['vms']).fill = fill
        ws.cell(row=row, column=4, value=a['ocpus']).fill = fill
        ws.cell(row=row, column=5, value=round(a['mem'], 1)).fill = fill
        ws.cell(row=row, column=7, value=total).font = group_font
        ws.cell(row=row, column=7).fill = fill
        row += 1

        # OCPU — use actual per-VM hours when custom_hours enabled
        if custom_hours and 'Hrs/Month' in a['df'].columns:
            ocpu_cost = float((a['df']['OCPUs'] * ocpu_rate * a['df']['Hrs/Month']).sum())
            mem_cost  = float((a['df']['Memory GB'] * mem_rate * a['df']['Hrs/Month']).sum())
            wos_cost  = float((a['df'].apply(
                lambda r: r['OCPUs'] * win_os_rate * r['Hrs/Month']
                          if (include_win_os and r['OS Family'] == 'Windows' and not r['BYOL']) else 0,
                axis=1)).sum())
            total_ocpu_hrs = int((a['df']['OCPUs'] * a['df']['Hrs/Month']).sum())
            wavg_hrs = round(total_ocpu_hrs / a['ocpus']) if a['ocpus'] else 0
            hrs_label = f"×variable hrs (OCPU-weighted avg: {wavg_hrs} hrs; total {total_ocpu_hrs:,} OCPU-hrs)"
        else:
            ocpu_cost = a['ocpus'] * ocpu_rate * HOURS_PER_MONTH
            mem_cost  = a['mem'] * mem_rate * HOURS_PER_MONTH
            wos_cost  = wos
            hrs_label = f"×{HOURS_PER_MONTH} hrs"

        for col, val in enumerate([SKU_E4_OCPU, f"Compute - E4 - OCPU (OCPU Per Hour) {hrs_label}",
                                    '', a['ocpus'], '', ocpu_rate, round(ocpu_cost, 2), "On-Demand"], 1):
            ws.cell(row=row, column=col, value=val).alignment = wrap
        row += 1

        # Memory
        for col, val in enumerate([SKU_E4_MEM, f"Compute - E4 - Memory (GB Per Hour) {hrs_label}",
                                    '', '', round(a['mem'], 1), mem_rate, round(mem_cost, 2), "On-Demand"], 1):
            ws.cell(row=row, column=col, value=val)
        row += 1

        # Block Storage
        stg_cost = a['stg'] * stg_rate
        for col, val in enumerate([SKU_BLOCK_STG, f"Block Volume - Storage ({round(a['stg'], 1)} GB)",
                                    '', '', round(a['stg'], 1), stg_rate, round(stg_cost, 2), "Balanced 10 VPU"], 1):
            ws.cell(row=row, column=col, value=val)
        row += 1

        # VPU
        vpu_cost = vpu_qty * vpu_rate
        for col, val in enumerate([SKU_BLOCK_VPU, f"Block Volume - Performance Units ({vpu_qty} VPU)",
                                    '', '', vpu_qty, vpu_rate, round(vpu_cost, 2), "10 VPU/GB (Balanced)"], 1):
            ws.cell(row=row, column=col, value=val)
        row += 1

        # Windows OS
        if include_win_os and wos_cost > 0:
            for col, val in enumerate(['B88318', f"Compute - Windows OS (OCPU Per Hour) {hrs_label}",
                                        '', a['ocpus'], '', win_os_rate, round(wos_cost, 2), "License Included"], 1):
                ws.cell(row=row, column=col, value=val)
            row += 1

        row += 1  # blank line

    if len(win_li) > 0:
        write_group(f"Windows VMs — License Included ({len(win_li)} VMs)", agg(win_li), True, light_grey)
    if len(win_byol) > 0:
        write_group(f"Windows VMs — BYOL ({len(win_byol)} VMs)", agg(win_byol), False, light_grey)
    if len(linux_df) > 0:
        write_group(f"Linux / Other VMs ({len(linux_df)} VMs)", agg(linux_df), False, light_grey)

    # 200 GB free tier discount row
    if free_tier_discount > 0:
        for col, val in enumerate(['', 'Block Volume Free Tier Discount (200 GB)', '', '', '', '', -round(free_tier_discount, 2), '200 GB/tenancy — applied once'], 1):
            ws.cell(row=row, column=col, value=val)
        row += 1

    # Grand total — recurring monthly
    net_total = round(grand_total - free_tier_discount, 2)
    for col, val in enumerate(['', 'GRAND TOTAL — Recurring Monthly (Compute + Storage)', '', '', '', '', net_total, ''], 1):
        c = ws.cell(row=row, column=col, value=val)
        c.font = bold
        c.fill = light_green
    row += 2

    # --- RackWare section ---
    if rw_params:
        ws.cell(row=row, column=1, value="RackWare Migration Cost (One-Time)").font = bold
        ws.cell(row=row, column=1).fill = light_blue
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        row += 1

        rw_ocpus   = rw_params['ocpus']
        rw_rate    = rw_params['rate']
        rw_cur     = rw_params['cur']
        rw_days    = rw_params['days']
        rw_hrs     = rw_params['hrs_day']
        rw_hours   = rw_days * rw_hrs
        rw_cost    = round(rw_ocpus * rw_rate * rw_hours, 2)

        params = [
            ("OCPUs to Migrate", rw_ocpus),
            (f"RackWare Rate ({rw_cur}/OCPU/hr)", f"{rw_cur} {rw_rate}"),
            ("Migration Duration (days)", rw_days),
            ("Hours/day", rw_hrs),
            ("Total Migration Hours", rw_hours),
        ]
        for label, val in params:
            ws.cell(row=row, column=2, value=label)
            ws.cell(row=row, column=3, value=val)
            row += 1

        for col, val in enumerate(['', f'RackWare Total (One-Time, {rw_cur})', '', '', '', '', f"{rw_cur} {rw_cost:,.2f}", 'PAYGO — OCI Marketplace'], 1):
            c = ws.cell(row=row, column=col, value=val)
            c.font = bold
        row += 2

    # --- Object Storage Backups section ---
    if obj_params:
        ws.cell(row=row, column=1, value="Object Storage — Backups (Monthly)").font = bold
        ws.cell(row=row, column=1).fill = light_blue
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        row += 1

        obj_gb      = obj_params['gb']
        obj_rate    = obj_params['rate']
        free_gb     = 20
        billable    = max(0, obj_gb - free_gb)
        obj_cost    = round(billable * obj_rate, 2)

        ws.cell(row=row, column=2, value="Estimated Backup Storage (GB)")
        ws.cell(row=row, column=3, value=obj_gb)
        row += 1
        ws.cell(row=row, column=2, value="Free Tier Deduction (GB)")
        ws.cell(row=row, column=3, value=free_gb)
        row += 1
        ws.cell(row=row, column=2, value="Billable Storage (GB)")
        ws.cell(row=row, column=3, value=billable)
        row += 1

        for col, val in enumerate([SKU_OBJ_STG, 'Object Storage - Standard (GB/month)', '', '', billable, obj_rate, obj_cost, 'First 20 GB free'], 1):
            c = ws.cell(row=row, column=col, value=val)
            c.font = bold
        row += 2

    # Quote note
    cell = ws.cell(row=row, column=1, value="Quote is for investment proposal only.")
    cell.font = red_bold
    row += 2

    # Disclaimer
    disc_cell = ws.cell(row=row, column=1, value=DISCLAIMER)
    disc_cell.alignment = Alignment(wrap_text=True, vertical="top")
    disc_cell.border = thin_border
    ws.merge_cells(start_row=row, start_column=1, end_row=row+4, end_column=9)
    ws.row_dimensions[row].height = 80

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()

if uploaded_file:
    try:
        df_vHost = pd.read_excel(uploaded_file, sheet_name='vHost')
        uploaded_file.seek(0)
        df_vInfo = pd.read_excel(uploaded_file, sheet_name='vInfo')
    except Exception as e:
        st.error(f"Failed to read file: {e}")
        st.stop()

    os_col = 'OS according to the configuration file'

    # --- vHost calculations ---
    vHost_filtered = df_vHost[df_vHost['CPU Model'] != 'VMware Virtual Processor'].copy()
    total_cores = int(vHost_filtered['# Cores'].sum())
    vHost_filtered['cores_used'] = (vHost_filtered['CPU usage %'] / 100) * vHost_filtered['# Cores']
    cores_based_on_avg_cpu = round(vHost_filtered['cores_used'].sum(), 2)
    avg_cpu_usage = round((cores_based_on_avg_cpu / total_cores) * 100, 2)
    total_memory_gb = round(vHost_filtered['# Memory'].sum() / 1024, 2)
    vHost_filtered['mem_used_gb'] = (vHost_filtered['Memory usage %'] / 100) * (vHost_filtered['# Memory'] / 1024)
    mem_based_on_avg_usage = round(vHost_filtered['mem_used_gb'].sum(), 2)
    avg_mem_usage = round((mem_based_on_avg_usage / total_memory_gb) * 100, 2)

    # --- All powered-on VMs ---
    powered_on_vms = df_vInfo[df_vInfo['Powerstate'] == 'poweredOn'].copy()
    powered_on_vms['Memory GB'] = round(powered_on_vms['Memory'] / 1024, 1)
    powered_on_vms['Provisioned GB'] = round(powered_on_vms['Provisioned MiB'] / 1024, 2)
    powered_on_vms['Provisioned TB'] = round(powered_on_vms['Provisioned MiB'] / 1024 / 1024, 3)
    powered_on_vms['OCPUs'] = (powered_on_vms['CPUs'] / 2).apply(lambda x: max(1, round(x)))
    powered_on_vms['OS Family'] = powered_on_vms[os_col].fillna('Unknown').apply(os_family)
    powered_on_vms['VMware Infra'] = (
        powered_on_vms['VM'].str.lower().str.startswith(VMWARE_NAME_PREFIXES) |
        powered_on_vms['VM'].str.lower().str.contains('vcsa', na=False) |
        powered_on_vms[os_col].str.contains('Photon CRX', case=False, na=False)
    )

    real_vms = powered_on_vms[~powered_on_vms['VMware Infra']].copy()
    vmware_vms = powered_on_vms[powered_on_vms['VMware Infra']].copy()

    tab1, tab2 = st.tabs(["📋  Infrastructure Summary", "📦  Migration BOM"])

    # ── TAB 1: Infrastructure Summary ─────────────────────
    with tab1:
        # OS breakdown data (needed for HTML below)
        real_os_counts = (
            real_vms[os_col].fillna('Unknown').value_counts().reset_index()
        )
        real_os_counts.columns = ['OS', 'Count']
        real_os_counts['Family'] = real_os_counts['OS'].apply(os_family)
        win_os_rows = real_os_counts[real_os_counts['Family'] == 'Windows']
        lin_os_rows = real_os_counts[real_os_counts['Family'] != 'Windows']

        def os_rows_html(df):
            if df.empty:
                return '<tr><td colspan="2" style="color:#aaa;font-style:italic;padding:6px 8px;">None</td></tr>'
            return ''.join(
                f'<tr><td style="padding:5px 8px;border-bottom:1px solid #eef0f4;color:#333;font-size:0.82rem;">{r.OS}</td>'
                f'<td style="padding:5px 8px;border-bottom:1px solid #eef0f4;color:#1a5c31;font-weight:700;text-align:right;">{r.Count}</td></tr>'
                for r in df.itertuples()
            )

        def stat(label, value, sub=None):
            sub_html = f'<div style="font-size:0.72rem;color:#999;margin-top:2px;">{sub}</div>' if sub else ''
            return f"""
            <div style="background:#fff;border:1px solid #dde2e8;border-radius:6px;
                        padding:16px 20px;box-shadow:0 1px 3px rgba(0,0,0,0.07);">
              <div style="font-size:0.68rem;font-weight:600;color:#888;
                          text-transform:uppercase;letter-spacing:0.06em;margin-bottom:6px;">{label}</div>
              <div style="font-size:1.6rem;font-weight:700;color:#1c1c1e;line-height:1.1;">{value}</div>
              {sub_html}
            </div>"""

        st.html(f"""
        <div style="padding:24px 28px 0;">

          <!-- Row 1: Host + CPU + Memory -->
          <div style="font-size:0.7rem;font-weight:700;color:#888;text-transform:uppercase;
                      letter-spacing:0.09em;margin-bottom:10px;">🖥️ &nbsp;Physical Infrastructure</div>
          <div style="display:grid;grid-template-columns:repeat(6,1fr);gap:12px;margin-bottom:24px;">
            {stat("Physical Hosts", len(vHost_filtered))}
            {stat("Total Cores", total_cores)}
            {stat("Total Memory", f"{total_memory_gb} GB")}
            {stat("Avg CPU Usage", f"{avg_cpu_usage}%", "weighted by host cores")}
            {stat("Actual Core Usage", cores_based_on_avg_cpu, "cores at avg utilisation")}
            {stat("Actual Mem Usage", f"{mem_based_on_avg_usage} GB", f"{avg_mem_usage}% of {total_memory_gb} GB")}
          </div>

          <!-- Row 2: VMs -->
          <div style="font-size:0.7rem;font-weight:700;color:#888;text-transform:uppercase;
                      letter-spacing:0.09em;margin-bottom:10px;">🗂️ &nbsp;Powered-On VMs</div>
          <div style="display:grid;grid-template-columns:repeat(3,1fr);gap:12px;margin-bottom:28px;">
            {stat("Total Powered-On VMs", len(powered_on_vms))}
            {stat("Workload VMs to Migrate", len(real_vms), "VMware infra excluded")}
            {stat("VMware Infra VMs (excluded)", len(vmware_vms), "vCLS, vRA, vCenter, Photon")}
          </div>

          <!-- Row 3: OS Breakdown side by side -->
          <div style="font-size:0.7rem;font-weight:700;color:#888;text-transform:uppercase;
                      letter-spacing:0.09em;margin-bottom:10px;">💿 &nbsp;OS Breakdown — Workload VMs</div>
          <div style="display:grid;grid-template-columns:1fr 1fr;gap:16px;margin-bottom:8px;">
            <div style="background:#fff;border:1px solid #dde2e8;border-radius:6px;
                        box-shadow:0 1px 3px rgba(0,0,0,0.07);overflow:hidden;">
              <div style="padding:10px 14px 8px;background:#f8f9fb;border-bottom:1px solid #eef0f4;
                          font-size:0.78rem;font-weight:700;color:#333;">🪟 &nbsp;Windows
                <span style="float:right;background:#1a5c31;color:#fff;border-radius:10px;
                             padding:1px 8px;font-size:0.7rem;">{len(win_os_rows)} versions</span>
              </div>
              <table style="width:100%;border-collapse:collapse;">
                <tr style="background:#f8f9fb;">
                  <th style="padding:5px 8px;text-align:left;font-size:0.68rem;color:#888;font-weight:600;text-transform:uppercase;">OS</th>
                  <th style="padding:5px 8px;text-align:right;font-size:0.68rem;color:#888;font-weight:600;text-transform:uppercase;">VMs</th>
                </tr>
                {os_rows_html(win_os_rows)}
              </table>
            </div>
            <div style="background:#fff;border:1px solid #dde2e8;border-radius:6px;
                        box-shadow:0 1px 3px rgba(0,0,0,0.07);overflow:hidden;">
              <div style="padding:10px 14px 8px;background:#f8f9fb;border-bottom:1px solid #eef0f4;
                          font-size:0.78rem;font-weight:700;color:#333;">🐧 &nbsp;Linux / Other
                <span style="float:right;background:#555;color:#fff;border-radius:10px;
                             padding:1px 8px;font-size:0.7rem;">{len(lin_os_rows)} versions</span>
              </div>
              <table style="width:100%;border-collapse:collapse;">
                <tr style="background:#f8f9fb;">
                  <th style="padding:5px 8px;text-align:left;font-size:0.68rem;color:#888;font-weight:600;text-transform:uppercase;">OS</th>
                  <th style="padding:5px 8px;text-align:right;font-size:0.68rem;color:#888;font-weight:600;text-transform:uppercase;">VMs</th>
                </tr>
                {os_rows_html(lin_os_rows)}
              </table>
            </div>
          </div>

        </div>
        """)

        with st.expander("VMware Infrastructure VMs (excluded from migration)"):
            st.caption("vCLS agents, vRealize agents, Photon CRX, vCenter — platform VMs, not customer workloads.")
            vmware_display = vmware_vms[['VM', os_col, 'CPUs', 'Memory GB']].rename(
                columns={os_col: 'OS', 'CPUs': 'vCPUs'}).reset_index(drop=True)
            st.dataframe(vmware_display, use_container_width=True, hide_index=True)

        with st.expander("Raw vHost Data"):
            st.dataframe(df_vHost, use_container_width=True)

    # ── TAB 2: Migration BOM ──────────────────────────────
    with tab2:

        def badge(n, color="#1a5c31"):
            return (f'<span style="display:inline-flex;align-items:center;justify-content:center;'
                    f'width:22px;height:22px;border-radius:50%;background:{color};color:#fff;'
                    f'font-weight:700;font-size:0.72rem;flex-shrink:0;'
                    f'box-shadow:0 1px 4px rgba(0,0,0,0.2);margin-right:6px;">{n}</span>')

        st.caption("Like-for-like migration based on provisioned values in VMware. VMware infra VMs excluded. 2 vCPU = 1 OCPU.")

        # ① ② Currency + shape inline
        col_cur, col_shape, _ = st.columns([1, 1, 2])
        with col_cur:
            st.markdown(f'{badge(1)} **Currency**', unsafe_allow_html=True)
            currency = st.selectbox("Currency", ["NZD", "USD", "AUD", "GBP", "EUR"], index=0, label_visibility="collapsed")
        with col_shape:
            st.markdown(f'{badge(2)} **OCI Shape**', unsafe_allow_html=True)
            shape = st.selectbox("OCI Shape", ["VM.Standard.E4.Flex", "VM.Standard.E5.Flex"], index=0, label_visibility="collapsed")

        # Fetch prices
        pricing_ok = False
        prices = {}
        with st.spinner("Fetching OCI pricing..."):
            try:
                prices = fetch_oci_prices(currency)
                ocpu_hr  = prices[SKU_E4_OCPU]
                mem_hr   = prices[SKU_E4_MEM]
                stg_mo   = prices[SKU_BLOCK_STG]
                vpu_mo   = prices[SKU_BLOCK_VPU]
                pricing_ok = True
            except Exception as e:
                st.error(f"Could not fetch OCI pricing: {e}")
                pricing_ok = False

        st.divider()

        bom_df = real_vms[['VM', 'OS Family', os_col, 'CPUs', 'OCPUs', 'Memory GB', 'Provisioned GB', 'Provisioned TB']].copy()
        bom_df = bom_df.rename(columns={os_col: 'OS', 'CPUs': 'vCPUs'})
        bom_df = bom_df.sort_values(['OS Family', 'VM']).reset_index(drop=True)

        # Migrate + BYOL toggles + per-VM hours
        bom_df['Migrate'] = True
        bom_df['BYOL'] = False
        bom_df['Hrs/Month'] = HOURS_PER_MONTH

        if pricing_ok:
            win_os_hr = prices.get('B88318', 0)

            st.markdown(f"**Shape:** `{shape}` &nbsp;|&nbsp; **Capacity:** On-Demand &nbsp;|&nbsp; **Storage:** Balanced (10 VPU) &nbsp;|&nbsp; **Prices:** Live OCI API ({currency})")
            st.info("⚠️ Quote is for investment proposal only.")
            st.html(f"""
            <div style="display:flex;gap:24px;align-items:center;
                        background:#fff;border:1px solid #dde2e8;border-radius:6px;
                        padding:10px 16px;margin:4px 0 8px;flex-wrap:wrap;">
              <div style="display:flex;align-items:center;gap:6px;font-size:0.8rem;color:#333;">
                {badge(3)} <span><strong>Migrate</strong> — uncheck to exclude a VM</span>
              </div>
              <div style="color:#dde2e8;">|</div>
              <div style="display:flex;align-items:center;gap:6px;font-size:0.8rem;color:#333;">
                {badge(4)} <span><strong>Hrs/Month</strong> — set actual run hours (compute only; storage always full month)</span>
              </div>
              <div style="color:#dde2e8;">|</div>
              <div style="display:flex;align-items:center;gap:6px;font-size:0.8rem;color:#333;">
                {badge(5)} <span><strong>BYOL</strong> — check if customer brings own Windows license</span>
              </div>
            </div>
            """)

            # Editable table
            edited = st.data_editor(
                bom_df[['Migrate', 'VM', 'OS Family', 'OS', 'vCPUs', 'OCPUs', 'Memory GB', 'Provisioned TB', 'Hrs/Month', 'BYOL']],
                column_config={
                    "Migrate": st.column_config.CheckboxColumn(
                        "③ Migrate",
                        help="Uncheck to exclude this VM from the BOM",
                        default=True,
                    ),
                    "Hrs/Month": st.column_config.NumberColumn(
                        "④ Hrs/Month",
                        help=f"Compute hours billed per month. Default {HOURS_PER_MONTH} (24×7). Storage always charged full month.",
                        min_value=1,
                        max_value=744,
                        step=1,
                    ),
                    "BYOL": st.column_config.CheckboxColumn(
                        "⑤ BYOL",
                        help="Check if customer brings their own Windows license",
                        default=False,
                    ),
                    "VM": st.column_config.TextColumn("VM", disabled=True),
                    "OS Family": st.column_config.TextColumn("OS Family", disabled=True),
                    "OS": st.column_config.TextColumn("OS", disabled=True),
                    "vCPUs": st.column_config.NumberColumn("vCPUs", disabled=True),
                    "OCPUs": st.column_config.NumberColumn("OCPUs", disabled=True),
                    "Memory GB": st.column_config.NumberColumn("Memory GB", disabled=True),
                    "Provisioned TB": st.column_config.NumberColumn("Provisioned TB", disabled=True),
                },
                use_container_width=True,
                hide_index=True,
                key="bom_editor"
            )

            # Merge selections back and filter to migrating VMs only
            bom_df['Migrate']   = edited['Migrate'].values
            bom_df['BYOL']      = edited['BYOL'].values
            bom_df['Hrs/Month'] = edited['Hrs/Month'].values
            excluded_count    = int((~bom_df['Migrate']).sum())
            bom_df            = bom_df[bom_df['Migrate']].copy()

            if excluded_count > 0:
                st.caption(f"ℹ️ {excluded_count} VM(s) excluded from BOM.")

            # Recalculate — compute uses per-VM hours, storage always full month
            def vm_cost(row):
                hrs     = int(row['Hrs/Month'])
                compute = (row['OCPUs'] * ocpu_hr + row['Memory GB'] * mem_hr) * hrs
                storage = row['Provisioned GB'] * stg_mo + BALANCED_VPU * row['Provisioned GB'] * vpu_mo
                win_os  = 0
                if row['OS Family'] == 'Windows' and not row['BYOL']:
                    win_os = row['OCPUs'] * win_os_hr * hrs
                return round(compute, 2), round(storage, 2), round(win_os, 2)

            bom_df[['Compute/mo', 'Storage/mo', 'WinOS/mo']] = bom_df.apply(
                lambda r: pd.Series(vm_cost(r)), axis=1
            )
            bom_df['Total/mo'] = round(bom_df['Compute/mo'] + bom_df['Storage/mo'] + bom_df['WinOS/mo'], 2)

            # 200 GB free tier always applied silently
            free_tier_discount = round(200 * stg_mo + 200 * BALANCED_VPU * vpu_mo, 2)
            total_monthly = round(bom_df['Total/mo'].sum() - free_tier_discount, 2)

            st.divider()

            win = bom_df[bom_df['OS Family'] == 'Windows']
            lin = bom_df[bom_df['OS Family'] != 'Windows']
            byol_label = "excl. Win License" if win['BYOL'].all() else ("incl. Win License" if not win['BYOL'].any() else "mixed BYOL")

            summary_html = f"""
<style>
.bs-wrap {{
  background: #fff;
  border: 1px solid #dde3ea;
  border-top: 3px solid #2d6a3f;
  border-radius: 4px;
  padding: 20px 24px 16px;
  margin: 4px 0 8px;
}}
.bs-table {{ width: 100%; border-collapse: collapse; font-family: -apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif; }}
.bs-table td {{ padding: 4px 12px 4px 0; vertical-align: top; width: 20%; }}
.bs-label {{ font-size: 0.7rem; color: #888; margin-bottom: 3px; text-transform: uppercase; letter-spacing: 0.05em; }}
.bs-value {{ font-size: 1.5rem; font-weight: 700; color: #1a1a1a; line-height: 1.15; }}
.bs-value.accent {{ color: #2d6a3f; }}
.bs-child .bs-value {{ font-size: 1.0rem; font-weight: 500; color: #444; }}
.bs-child .bs-label {{ padding-left: 14px; }}
.bs-child .bs-value {{ padding-left: 14px; }}
.bs-indent {{ color: #bbb; margin-right: 4px; }}
.bs-divider td {{ border-top: 1px solid #e8eaed; padding-top: 12px; }}
</style>
<div class="bs-wrap">
<table class="bs-table">
  <tr>
    <td><div class="bs-label">VMs to Migrate</div><div class="bs-value">{len(bom_df)}</div></td>
    <td><div class="bs-label">Total OCPUs</div><div class="bs-value">{int(bom_df['OCPUs'].sum())}</div></td>
    <td><div class="bs-label">Total RAM</div><div class="bs-value">{round(bom_df['Memory GB'].sum(), 1)} GB</div></td>
    <td><div class="bs-label">Total Storage</div><div class="bs-value">{round(bom_df['Provisioned TB'].sum(), 2)} TB</div></td>
    <td><div class="bs-label">Est. Monthly Cost ({currency})</div><div class="bs-value accent">{currency} {total_monthly:,.2f}</div></td>
  </tr>
  <tr class="bs-divider bs-child">
    <td><div class="bs-label"><span class="bs-indent">└</span> Windows ({byol_label})</div><div class="bs-value">{len(win)}</div></td>
    <td><div class="bs-label"><span class="bs-indent">└</span> Win OCPUs</div><div class="bs-value">{int(win['OCPUs'].sum())}</div></td>
    <td><div class="bs-label"><span class="bs-indent">└</span> Win RAM</div><div class="bs-value">{round(win['Memory GB'].sum(), 1)} GB</div></td>
    <td><div class="bs-label"><span class="bs-indent">└</span> Win Storage</div><div class="bs-value">{round(win['Provisioned TB'].sum(), 2)} TB</div></td>
    <td><div class="bs-label"><span class="bs-indent">└</span> Windows Cost ({currency})</div><div class="bs-value">{currency} {round(win['Total/mo'].sum(), 2):,.2f}</div></td>
  </tr>
  <tr class="bs-child">
    <td><div class="bs-label"><span class="bs-indent">└</span> Linux / Other</div><div class="bs-value">{len(lin)}</div></td>
    <td><div class="bs-label"><span class="bs-indent">└</span> Linux OCPUs</div><div class="bs-value">{int(lin['OCPUs'].sum())}</div></td>
    <td><div class="bs-label"><span class="bs-indent">└</span> Linux RAM</div><div class="bs-value">{round(lin['Memory GB'].sum(), 1)} GB</div></td>
    <td><div class="bs-label"><span class="bs-indent">└</span> Linux Storage</div><div class="bs-value">{round(lin['Provisioned TB'].sum(), 2)} TB</div></td>
    <td><div class="bs-label"><span class="bs-indent">└</span> Linux Cost ({currency})</div><div class="bs-value">{currency} {round(lin['Total/mo'].sum(), 2):,.2f}</div></td>
  </tr>
</table>
</div>
"""
            st.html(summary_html)

            st.divider()

            # Cost summary table with totals
            cost_cols = ['VM', 'OS Family', 'OCPUs', 'Memory GB', 'Provisioned TB', 'BYOL', 'Compute/mo', 'Storage/mo', 'WinOS/mo', 'Total/mo']
            cost_display = bom_df[cost_cols].copy()
            totals = pd.DataFrame([{
                'VM': 'TOTAL', 'OS Family': '', 'OCPUs': int(bom_df['OCPUs'].sum()),
                'Memory GB': round(bom_df['Memory GB'].sum(), 1),
                'Provisioned TB': round(bom_df['Provisioned TB'].sum(), 2),
                'BYOL': '',
                'Compute/mo': round(bom_df['Compute/mo'].sum(), 2),
                'Storage/mo': round(bom_df['Storage/mo'].sum(), 2),
                'WinOS/mo': round(bom_df['WinOS/mo'].sum(), 2),
                'Total/mo': round(bom_df['Total/mo'].sum(), 2),
            }])
            st.dataframe(pd.concat([cost_display, totals], ignore_index=True), use_container_width=True, hide_index=True)

            st.divider()

            st.html(f'<div style="display:flex;align-items:center;gap:6px;margin:20px 0 4px;">{badge(6)}<span style="font-size:0.9rem;font-weight:700;color:#1c1c1e;">🚚 RackWare Migration Cost (One-Time)</span></div>')
            st.html('RackWare on OCI Marketplace — PAYGO. Billed per OCPU per hour during migration window only. '
                    '<a href="https://marketplace.oracle.com/listings/migration-for-servers-vms/ocid1.mktpublisting.oc1.iad.amaaaaaao2ztryiax4zhb4csbzyv5dtoj3chyhxlw565qgx56pg6uokl7pea" '
                    'target="_blank" style="color:#1a5c31;font-weight:600;">For more info, click here ↗</a>')

            RW_RATES = {
                "USD": 2.53, "AUD": 3.99, "NZD": 4.3516, "EUR": 2.3529,
                "GBP": 1.8962, "CAD": 3.2462, "SGD": 3.4203, "CHF": 2.3215,
                "AED": 9.28, "SAR": 9.4875, "MYR": 12.0934, "ILS": 10.3224,
                "KRW": 3421.319, "BRL": 14.99, "MXN": 45.4641, "COP": 1073.44,
                "CLP": 2287.9, "DKK": 17.91, "PLN": 10.8, "CZK": 59.3,
                "HUF": 917.4286, "RON": 11.5874, "BGN": 4.7617, "BAM": 4.7612,
            }

            total_ocpus = int(bom_df['OCPUs'].sum())

            rw1, rw2, rw3, rw4 = st.columns(4)
            rw_ocpus   = rw1.number_input("OCPUs to Migrate", value=2, min_value=1, step=1,
                                          help=f"Total workload OCPUs: {total_ocpus}. Default 2 — adjust to actual migration batch size.")
            rw_cur     = rw2.selectbox("Rate Currency", list(RW_RATES.keys()), index=list(RW_RATES.keys()).index("USD"))
            rw_default = RW_RATES[rw_cur]
            rw_rate    = rw3.number_input(f"RackWare Rate ({rw_cur}/OCPU/hr)", value=rw_default, step=0.01, format="%.4f")
            rw_days    = rw4.number_input("Migration Duration (days)", value=7, min_value=1, step=1)
            rw_hrs_day = st.number_input("Hours/day RackWare runs", value=8, min_value=1, max_value=24, step=1)
            st.caption("💡 RackWare supports pause/resume and scheduled replication windows — it does not need to run 24/7. Typical migrations use 8 hrs/day (business hours). Adjust based on your agreed migration window with the customer.")

            rw_hours    = rw_days * rw_hrs_day
            rw_cost     = round(rw_ocpus * rw_rate * rw_hours, 2)

            st.caption(f"Formula: {rw_ocpus} OCPUs × {rw_cur} {rw_rate}/hr × {rw_hours} hrs ({rw_days} days × {rw_hrs_day} hrs/day)")

            rc1, rc2, rc3, rc4 = st.columns(4)
            rc1.metric("OCPUs Migrating", f"{rw_ocpus} of {total_ocpus}")
            rc2.metric("Migration Hours", rw_hours)
            rc3.metric(f"RackWare Cost ({rw_cur})", f"{rw_cur} {rw_cost:,.2f}")
            rc4.metric("Total BOM incl. RackWare (first month)", f"{rw_cur} {rw_cost:,.2f} + {currency} {total_monthly:,.2f}")

            st.divider()

            st.html(f'<div style="display:flex;align-items:center;gap:6px;margin:20px 0 4px;">{badge(7)}<span style="font-size:0.9rem;font-weight:700;color:#1c1c1e;">🗄️ Object Storage — Backups (Optional)</span></div>')
            st.caption("OCI Standard Object Storage for VM backups / snapshots post-migration. Customer estimate — adjust to actual backup retention needs.")

            obj_rate = prices.get(SKU_OBJ_STG, 0.0436101)
            os_col1, os_col2, os_col3 = st.columns(3)
            obj_gb = os_col1.number_input(
                "Estimated Backup Storage (GB)",
                value=1024,
                min_value=0,
                step=128,
                help="First 20 GB free. Enter total backup storage required across all VMs."
            )
            free_gb = 20
            billable_gb = max(0, obj_gb - free_gb)
            obj_cost_mo = round(billable_gb * obj_rate, 2)

            os_col2.metric("Billable GB (after 20 GB free)", f"{billable_gb:,} GB")
            os_col3.metric(f"Est. Monthly Cost ({currency})", f"{currency} {obj_cost_mo:,.2f}")

            st.caption(f"Formula: ({obj_gb:,} GB − {free_gb} GB free) × {currency} {obj_rate}/GB/mo = {currency} {obj_cost_mo:,.2f}/month  |  SKU: {SKU_OBJ_STG}")

            total_with_obj = round(total_monthly + obj_cost_mo, 2)
            if obj_cost_mo > 0:
                st.info(f"💡 Including object storage backups: **{currency} {total_with_obj:,.2f}/month** (compute + block storage + backups)")

        else:
            b1, b2, b3, b4 = st.columns(4)
            b1.metric("VMs to Migrate", len(bom_df))
            b2.metric("Total OCPUs", int(bom_df['OCPUs'].sum()))
            b3.metric("Total RAM", f"{round(bom_df['Memory GB'].sum(), 1)} GB")
            b4.metric("Total Storage", f"{round(bom_df['Provisioned TB'].sum(), 2)} TB")

            totals = pd.DataFrame([{
                'VM': 'TOTAL', 'OS Family': '', 'OS': '',
                'vCPUs': int(bom_df['vCPUs'].sum()),
                'OCPUs': int(bom_df['OCPUs'].sum()),
                'Memory GB': round(bom_df['Memory GB'].sum(), 1),
                'Provisioned GB': round(bom_df['Provisioned GB'].sum(), 1),
                'Provisioned TB': round(bom_df['Provisioned TB'].sum(), 2),
            }])
            bom_display = pd.concat([bom_df, totals], ignore_index=True)
            st.dataframe(bom_display, use_container_width=True, hide_index=True)

        try:
            if pricing_ok and prices:
                _rw_params  = dict(ocpus=rw_ocpus, rate=rw_rate, cur=rw_cur, days=rw_days, hrs_day=rw_hrs_day) if pricing_ok else None
                _obj_params = dict(gb=obj_gb, rate=obj_rate) if pricing_ok else None
                excel_data  = to_oci_excel(bom_df, prices, currency, shape, _rw_params, _obj_params, custom_hours=True, free_tier_discount=free_tier_discount)
            else:
                excel_data = to_excel(bom_df)
        except Exception:
            excel_data = to_excel(bom_df)
        st.html(f'<div style="display:flex;align-items:center;gap:6px;margin:20px 0 6px;">{badge(8, "#e8a800")}<span style="font-size:0.9rem;font-weight:700;color:#1c1c1e;">Download BOM — OCI Investment Proposal format</span></div>')
        st.download_button(
            label="⬇️ Export BOM to Excel",
            data=excel_data,
            file_name="migration_bom.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
